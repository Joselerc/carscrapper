"""
Exportadores para guardar datos de scraping en diferentes formatos
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Any
import pandas as pd

from .models import NormalizedListing, SearchResult

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exportador para guardar resultados de scraping en formato Excel
    """
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_listings(
        self,
        listings: List[NormalizedListing],
        filename: Optional[str] = None,
        include_metadata: bool = True
    ) -> str:
        """
        Exporta una lista de listings a Excel
        
        Args:
            listings: Lista de listings normalizados
            filename: Nombre del archivo (opcional, se genera automáticamente si no se proporciona)
            include_metadata: Si incluir metadatos en el Excel
            
        Returns:
            Ruta del archivo generado
        """
        if not listings:
            raise ValueError("No hay listings para exportar")
        
        # Generar nombre de archivo si no se proporciona
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            sources = set(listing.source for listing in listings)
            source_str = "_".join(sorted(sources))
            filename = f"car_listings_{source_str}_{timestamp}.xlsx"
        
        # Asegurar extensión .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_dir / filename
        
        # Convertir listings a DataFrame
        df = self._listings_to_dataframe(listings, include_metadata)
        
        # Guardar en Excel con formato
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Listings', index=False)
            
            # Formatear la hoja
            worksheet = writer.sheets['Listings']
            self._format_worksheet(worksheet, df)
        
        logger.info(f"Exportados {len(listings)} listings a {filepath}")
        return str(filepath)
    
    def export_search_results(
        self,
        results: List[SearchResult],
        filename: Optional[str] = None,
        include_metadata: bool = True
    ) -> str:
        """
        Exporta múltiples SearchResult a Excel
        
        Args:
            results: Lista de resultados de búsqueda
            filename: Nombre del archivo
            include_metadata: Si incluir metadatos
            
        Returns:
            Ruta del archivo generado
        """
        all_listings = []
        for result in results:
            all_listings.extend(result.listings)
        
        return self.export_listings(all_listings, filename, include_metadata)
    
    def _listings_to_dataframe(self, listings: List[NormalizedListing], include_metadata: bool) -> pd.DataFrame:
        """Convierte listings a DataFrame con campos unificados"""
        
        data = []
        for listing in listings:
            row = {
                # Identificación
                'listing_id': listing.listing_id,
                'source': listing.source,
                'url': listing.url,
                'scraped_at': listing.scraped_at.replace(tzinfo=None) if listing.scraped_at else None,
                
                # Información básica del vehículo
                'title': listing.title,
                'make': listing.make,
                'model': listing.model,
                
                # Precios (campos unificados)
                'price_gross_eur': listing.price_eur,  # Precio bruto
                'price_net_eur': listing.price_net_eur,  # Precio neto
                'original_price': listing.price_original.amount if listing.price_original else None,
                'original_currency': listing.price_original.currency_code if listing.price_original else None,
                
                # Características técnicas
                'year': listing.first_registration.year if listing.first_registration else None,
                'month': listing.first_registration.month if listing.first_registration else None,
                'mileage_km': listing.mileage_km,
                'power_hp': listing.power_hp,
                'power_kw': listing.power_kw,
                'engine_displacement_cc': listing.engine_displacement_cc,
                
                # Especificaciones
                'fuel_type': listing.fuel_type,
                'transmission': listing.transmission,
                'body_type': listing.body_type,
                'doors': listing.doors,
                'seats': listing.seats,
                'color_exterior': listing.color_exterior,
                
                # Emisiones y consumo
                'co2_emissions_g_km': listing.co2_emissions_g_km,
                'consumption_combined_l_100km': listing.consumption_l_100km.combined if listing.consumption_l_100km else None,
                'consumption_urban_l_100km': listing.consumption_l_100km.urban if listing.consumption_l_100km else None,
                'consumption_highway_l_100km': listing.consumption_l_100km.highway if listing.consumption_l_100km else None,
                
                # Ubicación
                'country_code': listing.location.country_code if listing.location else None,
                'region': listing.location.region if listing.location else None,
                'province': listing.location.province if listing.location else None,
                'city': listing.location.city if listing.location else None,
                'postal_code': listing.location.postal_code if listing.location else None,
                'latitude': listing.location.latitude if listing.location else None,
                'longitude': listing.location.longitude if listing.location else None,
                
                # Vendedor
                'seller_type': listing.seller.type if listing.seller else None,
                'seller_name': listing.seller.name if listing.seller else None,
                'seller_rating': listing.seller.rating if listing.seller else None,
                'seller_rating_count': listing.seller.rating_count if listing.seller else None,
                'seller_phone': listing.seller.phone if listing.seller else None,
                'seller_email': listing.seller.email if listing.seller else None,
                'seller_vat_number': listing.seller.vat_number if listing.seller else None,
                'seller_dealer_id': listing.seller.dealer_id if listing.seller else None,
                
                # Descripción
                'description': listing.description,
            }
            
            # Añadir metadatos si se solicita
            if include_metadata and listing.metadata:
                row.update({
                    'advert_type': listing.metadata.advert_type,
                    'vehicle_id': listing.metadata.vehicle_id,
                    'environment_badge': listing.metadata.environment_badge,
                    'hsn_tsn': listing.metadata.hsn_tsn,
                    'certified': listing.metadata.certified,
                    'publish_date': listing.metadata.publish_date.replace(tzinfo=None) if listing.metadata.publish_date else None,
                    'update_date': listing.metadata.update_date.replace(tzinfo=None) if listing.metadata.update_date else None,
                    'exportable': listing.metadata.exportable,
                })
            
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """Aplica formato a la hoja de Excel"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Limitar el ancho máximo
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        worksheet.freeze_panes = "A2"


class CSVExporter:
    """
    Exportador simple para CSV
    """
    
    def __init__(self, output_dir: str = "exports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
    
    def export_listings(
        self,
        listings: List[NormalizedListing],
        filename: Optional[str] = None
    ) -> str:
        """Exporta listings a CSV"""
        
        if not listings:
            raise ValueError("No hay listings para exportar")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            sources = set(listing.source for listing in listings)
            source_str = "_".join(sorted(sources))
            filename = f"car_listings_{source_str}_{timestamp}.csv"
        
        if not filename.endswith('.csv'):
            filename += '.csv'
        
        filepath = self.output_dir / filename
        
        # Usar el mismo método de conversión que Excel
        exporter = ExcelExporter()
        df = exporter._listings_to_dataframe(listings, include_metadata=True)
        
        # Guardar como CSV
        df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        logger.info(f"Exportados {len(listings)} listings a {filepath}")
        return str(filepath)


__all__ = ["ExcelExporter", "CSVExporter"]
